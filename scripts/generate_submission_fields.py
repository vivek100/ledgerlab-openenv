#!/usr/bin/env python3
"""
Generate submission_fields for each task in task_manifest.json by
inspecting the gold deliverable Excel files with openpyxl.

For each gold file, extracts 3-5 key numeric/structural facts
(totals, row counts, distinct categories, etc.) and writes them
back into the manifest as submission_fields.
"""

import json
from pathlib import Path
from typing import Any

import openpyxl


BASE_DIR = Path(__file__).resolve().parent.parent
MANIFEST_PATH = BASE_DIR / "data" / "tasks" / "task_manifest.json"
WORKSPACES_DIR = BASE_DIR / "data" / "tasks" / "workspaces"


def round_if_float(val: Any, decimals: int = 2) -> Any:
    if isinstance(val, float):
        rounded = round(val, decimals)
        if rounded == int(rounded):
            return int(rounded)
        return rounded
    return val


def field(key: str, typ: str, desc: str, expected: Any, tolerance: float | None = None) -> dict:
    f = {
        "key": key,
        "type": typ,
        "description": desc,
        "expected": round_if_float(expected),
    }
    if tolerance is not None:
        f["tolerance"] = tolerance
    return f


def count_data_rows(ws, header_row: int = 1) -> int:
    """Count non-empty rows after the header row."""
    count = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if any(v is not None for v in row):
            count += 1
    return count


def find_row_values(ws, label: str) -> list[tuple[str, Any]]:
    """Find a row where any cell matches label (case-insensitive, stripped).
    Returns list of (col_letter, value) for non-None cells."""
    label_lower = label.lower().strip()
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.lower().strip() == label_lower:
                result = []
                for c in row:
                    if c.value is not None:
                        col_letter = openpyxl.utils.get_column_letter(c.column)
                        result.append((col_letter, c.value))
                return result
    return []


def get_col(row_vals: list[tuple[str, Any]], col: str, numeric: bool = True) -> Any:
    """Extract value from a specific column letter in find_row_values output."""
    for col_letter, val in row_vals:
        if col_letter == col:
            if numeric and isinstance(val, (int, float)):
                return val
            if not numeric:
                return val
    return None


def distinct_column_values(ws, col_idx: int, header_row: int = 1) -> set:
    """Get distinct non-empty values in a column (1-indexed) after the header row."""
    vals = set()
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if col_idx - 1 < len(row) and row[col_idx - 1] is not None:
            vals.add(str(row[col_idx - 1]).strip())
    return vals


# ── Task-specific extractors ──────────────────────────────────────────

def extract_c357f0e2(wb: openpyxl.Workbook) -> list[dict]:
    """UAT Plan.xlsx - test case spreadsheet."""
    ws = wb["UAT Test Plan"]
    data_rows = count_data_rows(ws, header_row=2)
    roles = distinct_column_values(ws, col_idx=2, header_row=2)
    roles.discard("Role")
    modules = distinct_column_values(ws, col_idx=3, header_row=2)
    modules.discard("Module")
    return [
        field("sheet_count", "integer", "How many sheets does the workbook have?", len(wb.sheetnames)),
        field("total_test_cases", "integer", "How many test case data rows are there (excluding header rows)?", data_rows),
        field("distinct_roles", "integer", "How many distinct Role values appear in the test cases?", len(roles)),
        field("distinct_modules", "integer", "How many distinct Module values appear in the test cases?", len(modules)),
    ]


def extract_f841ddcf(wb: openpyxl.Workbook) -> list[dict]:
    """PO Log June Ships.xlsx - purchase order tracking."""
    ws_summary = wb["Summary Table"]
    grand_total_row = find_row_values(ws_summary, "Grand Total")
    order_val = get_col(grand_total_row, "C")
    shipped_val = get_col(grand_total_row, "D")

    ws_log = wb["PO Log"]
    log_rows = 0
    for row in ws_log.iter_rows(min_row=3, values_only=True):
        if any(v is not None for v in row):
            log_rows += 1

    return [
        field("sheet_count", "integer", "How many sheets does the workbook have?", len(wb.sheetnames)),
        field("po_log_data_rows", "integer", "How many data rows are in the PO Log sheet (excluding headers)?", log_rows),
        field("summary_grand_total_order_value", "number", "What is the Grand Total 'Sum of Order Value $ Cost' on the Summary Table?", order_val, tolerance=1.0),
        field("summary_grand_total_shipped_value", "number", "What is the Grand Total 'Sum of Shipped Value $ Cost' on the Summary Table?", shipped_val, tolerance=1.0),
    ]


def extract_7b08cd4d(wb: openpyxl.Workbook) -> list[dict]:
    """Fall Music Tour Output.xlsx - P&L tour financials."""
    ws = wb["P&L Tour"]
    gross_rev_row = find_row_values(ws, "Total Gross Revenue")
    net_rev_row = find_row_values(ws, "Total Net Revenue")
    expenses_row = find_row_values(ws, "Total Expenses")
    net_income_row = find_row_values(ws, "Net Income")

    return [
        field("total_gross_revenue", "number", "What is the Total Gross Revenue (Total column)?", get_col(gross_rev_row, "G"), tolerance=100),
        field("total_net_revenue", "number", "What is the Total Net Revenue (Total column)?", get_col(net_rev_row, "G"), tolerance=100),
        field("total_expenses", "number", "What is the Total Expenses (Total column)?", get_col(expenses_row, "G"), tolerance=100),
        field("net_income", "number", "What is the Net Income (Total column)?", get_col(net_income_row, "G"), tolerance=100),
    ]


def extract_327fbc21(wb: openpyxl.Workbook) -> list[dict]:
    """May Sales Plan final.xlsx - store-level sales forecasts."""
    ws = wb["May Plan"]
    total_row = find_row_values(ws, "TOTAL STORES")
    comp_row = find_row_values(ws, "COMP STORES")
    closed_row = find_row_values(ws, "CLOSED STORES")

    return [
        field("total_stores_plan_total", "number", "What is the TOTAL STORES grand total (rightmost summary column)?", get_col(total_row, "I"), tolerance=50),
        field("comp_stores_plan_total", "number", "What is the COMP STORES grand total (rightmost summary column)?", get_col(comp_row, "I"), tolerance=50),
        field("closed_stores_plan_total", "number", "What is the CLOSED STORES grand total (rightmost summary column)?", get_col(closed_row, "I"), tolerance=50),
    ]


def extract_9e39df84(wb: openpyxl.Workbook) -> list[dict]:
    """Dashboard Output.xlsx - operator/machine output dashboard."""
    ws = wb["Dashboard"]
    grand_total_output = None
    avg_output = None
    for row in ws.iter_rows(values_only=True):
        vals = list(row)
        # V=col22, W=col23, Y=col25, Z=col26 (1-indexed)
        v_idx, w_idx, y_idx, z_idx = 21, 22, 24, 25
        v_val = vals[v_idx] if len(vals) > v_idx else None
        w_val = vals[w_idx] if len(vals) > w_idx else None
        y_val = vals[y_idx] if len(vals) > y_idx else None
        z_val = vals[z_idx] if len(vals) > z_idx else None
        if isinstance(v_val, str) and v_val.lower().strip() == "grand total":
            if isinstance(w_val, (int, float)):
                grand_total_output = w_val
        if isinstance(y_val, str) and y_val.lower().strip() == "grand total":
            if isinstance(z_val, (int, float)):
                avg_output = z_val

    op_sheet_name = [s for s in wb.sheetnames if "operator" in s.lower()][0]
    ws_op = wb[op_sheet_name]
    op_rows = 0
    for row in ws_op.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            op_rows += 1

    return [
        field("sheet_count", "integer", "How many sheets does the workbook have?", len(wb.sheetnames)),
        field("operator_data_rows", "integer", "How many data rows are in the 'Operator Output Data' sheet?", op_rows),
        field("grand_total_output", "number", "What is the Grand Total 'Sum of Total Output' on the Dashboard?", grand_total_output, tolerance=10),
        field("average_output_per_operator", "number", "What is the Grand Total 'Average of Average Output' on the Dashboard?", avg_output, tolerance=5),
    ]


def extract_19403010(wb: openpyxl.Workbook) -> list[dict]:
    """XR Retailer 2023 Sales Performance Analysis."""
    ws_top = wb["TOP LINE ANALYSIS"]
    total_row = find_row_values(ws_top, "TOTAL")
    sales_2023 = get_col(total_row, "C")
    sales_2022 = get_col(total_row, "D")

    return [
        field("sheet_count", "integer", "How many sheets does the workbook have?", len(wb.sheetnames)),
        field("total_xr_sales_fy2023", "number", "What is the TOTAL XR Sales Dollars FY 2023 (in 000s) on the TOP LINE ANALYSIS sheet?", sales_2023, tolerance=10),
        field("total_xr_sales_fy2022", "number", "What is the TOTAL XR Sales Dollars FY 2022 LY (in 000s) on the TOP LINE ANALYSIS sheet?", sales_2022, tolerance=10),
        field("pct_change_ty_vs_ly", "number", "What is the TOTAL % CHG TY vs LY on the TOP LINE ANALYSIS?",
              round((sales_2023 - sales_2022) / sales_2022, 4) if sales_2023 and sales_2022 else None, tolerance=0.01),
    ]


def extract_47ef842d(wb: openpyxl.Workbook) -> list[dict]:
    """Inventory final.xlsx - SKU inventory snapshot."""
    ws = wb["Summary"]
    data_rows = count_data_rows(ws, header_row=4)
    return [
        field("sheet_count", "integer", "How many sheets does the workbook have?", len(wb.sheetnames)),
        field("total_skus", "integer", "How many SKU data rows are in the Summary sheet?", data_rows),
    ]


def extract_76418a2c(wb: openpyxl.Workbook) -> list[dict]:
    """Daily Shipment Manifest 062525.xlsx."""
    ws = wb["Sheet1"]
    shipment_rows = count_data_rows(ws, header_row=2)
    total_weight = 0.0
    total_cost = 0.0
    for row in ws.iter_rows(min_row=3, values_only=True):
        if len(row) > 2 and isinstance(row[2], (int, float)):
            total_weight += row[2]
        if len(row) > 6 and isinstance(row[6], (int, float)):
            total_cost += row[6]

    return [
        field("total_shipments", "integer", "How many shipment rows are in the manifest?", shipment_rows),
        field("total_weight_lbs", "number", "What is the total weight in lbs across all shipments?", total_weight, tolerance=1),
        field("total_shipment_cost", "number", "What is the total shipment cost across all shipments?", total_cost, tolerance=1),
    ]


def extract_7ed932dd(wb: openpyxl.Workbook) -> list[dict]:
    """Additional Shipments Needed Updated - July 2025."""
    ws = wb["Current Inventory"]
    data_rows = 0
    for row in ws.iter_rows(min_row=4, values_only=True):
        if any(v is not None for v in row):
            if row[0] and str(row[0]).startswith("PRD"):
                data_rows += 1
    remaining_days = None
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        if len(row) >= 5 and isinstance(row[4], (int, float)):
            remaining_days = int(row[4])

    return [
        field("sheet_count", "integer", "How many sheets does the workbook have?", len(wb.sheetnames)),
        field("total_products", "integer", "How many product (PRD-xxx) rows are listed?", data_rows),
        field("remaining_days_in_month", "integer", "What is the 'Remaining Days in the Month' value?", remaining_days),
    ]


def extract_b7a5912e(wb: openpyxl.Workbook) -> list[dict]:
    """Daily Closed Operational Report June 27, 2025."""
    ws = wb["Sheet1"]

    total_rentals = get_col(find_row_values(ws, "Total Closed Rentals"), "C")
    total_days = get_col(find_row_values(ws, "Total Rental Days"), "C")
    total_revenue = get_col(find_row_values(ws, "Total Revenue"), "C")
    avg_daily_rate = get_col(find_row_values(ws, "Average Daily Rate"), "C")

    categories = 0
    for row in ws.iter_rows(min_row=16, max_row=23, values_only=True):
        if row and row[0] and isinstance(row[0], str) and row[0].strip():
            categories += 1

    return [
        field("total_closed_rentals", "integer", "What is the Total Closed Rentals count?", int(total_rentals) if total_rentals else None),
        field("total_rental_days", "integer", "What is the Total Rental Days?", int(total_days) if total_days else None),
        field("total_revenue", "number", "What is the Total Revenue?", total_revenue, tolerance=10),
        field("vehicle_category_count", "integer", "How many vehicle categories are in the breakdown table?", categories),
        field("average_daily_rate", "number", "What is the Average Daily Rate?", avg_daily_rate, tolerance=1),
    ]


def extract_d7cfae6f(wb: openpyxl.Workbook) -> list[dict]:
    """Beutist Q124 Risks SET SELLING final_v2.xlsx."""
    ws = wb["Set Sales 2023"]
    grand_row = find_row_values(ws, "GRAND TOTAL")
    skincare_row = find_row_values(ws, "SKIN CARE Total")
    makeup_row = find_row_values(ws, "MAKEUP Total")
    fragrance_row = find_row_values(ws, "FRAGRANCE Total")

    return [
        field("grand_total_ytd_set_sales", "number", "What is the GRAND TOTAL YTD SET SALES?", get_col(grand_row, "C"), tolerance=10),
        field("skincare_total_ytd", "number", "What is the SKIN CARE Total YTD SET SALES?", get_col(skincare_row, "C"), tolerance=5),
        field("makeup_total_ytd", "number", "What is the MAKEUP Total YTD SET SALES?", get_col(makeup_row, "C"), tolerance=5),
        field("fragrance_total_ytd", "number", "What is the FRAGRANCE Total YTD SET SALES?", get_col(fragrance_row, "C"), tolerance=5),
    ]


def extract_efca245f(wb: openpyxl.Workbook) -> list[dict]:
    """Running Board Recovery Plan Analysis.xlsx."""
    sheets = wb.sheetnames
    rows_per_sheet = {}
    for sn in sheets:
        ws = wb[sn]
        count = 0
        for row in ws.iter_rows(min_row=4, values_only=True):
            if any(v is not None for v in row):
                count += 1
        rows_per_sheet[sn] = count

    return [
        field("sheet_count", "integer", "How many sheets (scenarios) does the workbook have?", len(sheets)),
        field("sheet_names", "text", "What are the sheet names (comma-separated)?", ", ".join(sheets)),
        field("data_rows_sheet1", "integer", f"How many data rows are in the '{sheets[0]}' sheet?", rows_per_sheet[sheets[0]]),
    ]


EXTRACTORS = {
    "c357f0e2": extract_c357f0e2,
    "f841ddcf": extract_f841ddcf,
    "7b08cd4d": extract_7b08cd4d,
    "327fbc21": extract_327fbc21,
    "9e39df84": extract_9e39df84,
    "19403010": extract_19403010,
    "47ef842d": extract_47ef842d,
    "76418a2c": extract_76418a2c,
    "7ed932dd": extract_7ed932dd,
    "b7a5912e": extract_b7a5912e,
    "d7cfae6f": extract_d7cfae6f,
    "efca245f": extract_efca245f,
}


def generic_extract(wb: openpyxl.Workbook) -> list[dict]:
    """Fallback: extract sheet count and row counts."""
    fields = [
        field("sheet_count", "integer", "How many sheets does the workbook have?", len(wb.sheetnames)),
        field("sheet_names", "text", "What are the sheet names (comma-separated)?", ", ".join(wb.sheetnames)),
    ]
    for sn in wb.sheetnames[:3]:
        ws = wb[sn]
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            if any(v is not None for v in row):
                row_count += 1
        fields.append(field(
            f"data_rows_{sn.lower().replace(' ', '_')[:30]}",
            "integer",
            f"How many non-empty rows are in the '{sn}' sheet?",
            row_count,
        ))
    return fields


def main():
    with open(MANIFEST_PATH) as f:
        manifest = json.load(f)

    print(f"Loaded manifest with {len(manifest)} tasks\n")

    for task in manifest:
        task_id = task["task_id"]
        short_id = task_id[:8]
        deliverables = task.get("expected_deliverables", [])
        if not deliverables:
            print(f"[{short_id}] No expected deliverables, skipping.")
            continue

        gold_file = deliverables[0]
        gold_path = WORKSPACES_DIR / task_id / "gold" / gold_file

        if not gold_path.exists():
            print(f"[{short_id}] Gold file not found: {gold_path}")
            continue

        print(f"[{short_id}] Processing: {gold_file}")
        try:
            wb = openpyxl.load_workbook(str(gold_path), data_only=True, read_only=True)

            extractor = EXTRACTORS.get(short_id, generic_extract)
            fields = extractor(wb)
            wb.close()

            task["submission_fields"] = fields
            print(f"  -> Generated {len(fields)} fields:")
            for fld in fields:
                exp = fld['expected']
                if isinstance(exp, str) and len(exp) > 60:
                    exp = exp[:60] + "..."
                tol = f" (±{fld['tolerance']})" if 'tolerance' in fld else ""
                print(f"     {fld['key']:40s} = {exp}{tol}")
            print()

        except Exception as e:
            print(f"  ERROR: {e}")
            import traceback
            traceback.print_exc()
            print()

    with open(MANIFEST_PATH, "w") as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False)

    print("=" * 70)
    print(f"Wrote updated manifest to {MANIFEST_PATH}")
    tasks_with_fields = sum(1 for t in manifest if "submission_fields" in t)
    print(f"Tasks with submission_fields: {tasks_with_fields}/{len(manifest)}")


if __name__ == "__main__":
    main()
