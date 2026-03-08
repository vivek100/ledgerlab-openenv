#!/usr/bin/env python3
"""
Generate submission_fields for tasks using:
1) LLM extraction (optional), and
2) deterministic workbook heuristics fallback.

This script is designed to process tasks one-by-one and always produce 2-4
question-style fields with expected answers for reward scoring.
"""

import argparse
import json
import os
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent
DEFAULT_MANIFEST = BASE_DIR / "data" / "tasks" / "task_manifest.json"
WORKSPACES_DIR = BASE_DIR / "data" / "tasks" / "workspaces"
GENERIC_KEYS = {"sheet_count", "sheet_names"}


def is_generic_fields(fields: List[Dict[str, Any]]) -> bool:
    if not fields:
        return False
    for f in fields:
        key = str(f.get("key", ""))
        if key in GENERIC_KEYS or key.startswith("data_rows_"):
            continue
        return False
    return True


def to_number(v: Any) -> Optional[float]:
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        if not s:
            return None
        try:
            return float(s)
        except Exception:
            return None
    return None


def snake(s: str) -> str:
    s = s.lower().strip()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s or "value"


def round_value(v: Any) -> Any:
    if isinstance(v, float):
        r = round(v, 4)
        if abs(r - int(r)) < 1e-9:
            return int(r)
        return r
    return v


def tolerance_for_value(v: float) -> float:
    av = abs(v)
    if av < 1000:
        return 1.0
    if av < 10000:
        return 10.0
    return 100.0


def sample_sheet(ws, max_rows: int = 12, max_cols: int = 12) -> List[List[str]]:
    out: List[List[str]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        cells = []
        for c in list(row)[:max_cols]:
            if c is None:
                cells.append("")
            else:
                cells.append(str(c))
        out.append(cells)
    return out



def field(
    key: str,
    field_type: str,
    description: str,
    expected: Any,
    tolerance: Optional[float] = None,
) -> Dict[str, Any]:
    item = {
        "key": snake(key)[:60],
        "type": field_type,
        "description": description,
        "expected": round_value(expected),
    }
    if field_type == "number":
        item["tolerance"] = tolerance if tolerance is not None else tolerance_for_value(float(expected))
    return item


def find_row_by_first_value(ws, first_value: str) -> Optional[List[Any]]:
    target = first_value.strip().lower()
    for row in ws.iter_rows(values_only=True):
        first = row[0] if row else None
        if isinstance(first, str) and first.strip().lower() == target:
            return list(row)
    return None


def extract_efca245f(task: Dict[str, Any], wb: openpyxl.Workbook) -> List[Dict[str, Any]]:
    current = wb["Current Capacity and Cells"]
    relocated = wb["Relocated Grill Guard"]
    ten_hour = wb["10 hr Shift Relocate Grill Guar"]
    current_last = [current.cell(108, c).value for c in range(1, 12)]
    relocated_last = [relocated.cell(108, c).value for c in range(1, 12)]
    ten_hour_last = [ten_hour.cell(108, c).value for c in range(1, 12)]
    return [
        field("scenario_count", "integer", "How many scenario sheets are included in the workbook?", len(wb.sheetnames)),
        field("current_capacity_final_crew_cab_open_po", "integer", "What is the final Crew Cab cumulative open purchase orders value in 'Current Capacity and Cells'?", int(current_last[4] or 0)),
        field("current_capacity_final_extended_cab_open_po", "integer", "What is the final Extended Cab cumulative open purchase orders value in 'Current Capacity and Cells'?", int(current_last[7] or 0)),
        field("relocated_grill_guard_final_extended_cab_open_po", "integer", "What is the final Extended Cab cumulative open purchase orders value in 'Relocated Grill Guard'?", int(relocated_last[7] or 0)),
    ]


def extract_83d10b06(task: Dict[str, Any], wb: openpyxl.Workbook) -> List[Dict[str, Any]]:
    ws = wb["Sample Size"]
    return [
        field("population_dataset_size", "integer", "What is the dataset size listed on the 'Sample Size' sheet?", int(ws.cell(1, 3).value)),
        field("audit_sample_size", "integer", "What is the final sample size listed on the 'Sample Size' sheet?", int(ws.cell(5, 3).value)),
        field("sample_methodology", "text", "What confidence level methodology is listed on the 'Sample Size' sheet?", str(ws.cell(2, 3).value).strip()),
        field("error_rate_text", "text", "What error rate text is listed on the 'Sample Size' sheet?", str(ws.cell(3, 3).value).strip()),
    ]


def extract_7bbfcfe9(task: Dict[str, Any], wb: openpyxl.Workbook) -> List[Dict[str, Any]]:
    rights = wb["Exercise of Rights"]
    interest = wb["Maximum Interest Rate"]
    rights_rows = list(rights.iter_rows(values_only=True))[1:]
    interest_rows = list(interest.iter_rows(values_only=True))[1:]
    mir_rows = [r for r in interest_rows if isinstance(r[1], str) and r[1].startswith("MIR-")]
    return [
        field("exercise_of_rights_question_count", "integer", "How many Exercise of Rights questions are listed?", len(rights_rows)),
        field("maximum_interest_rate_question_count", "integer", "How many MIR-specific Maximum Interest Rate questions are listed?", len(mir_rows)),
        field("total_specific_scra_question_count", "integer", "How many specific SCRA questions are listed across both sheets, excluding the leading question?", len(rights_rows) + len(mir_rows)),
        field("first_exercise_of_rights_id", "text", "What is the first Exercise of Rights question ID?", str(rights_rows[0][1]).strip()),
    ]


def extract_dfb4e0cd(task: Dict[str, Any], wb: openpyxl.Workbook) -> List[Dict[str, Any]]:
    from collections import Counter

    ws = wb["Final Analysis"]
    rows = [list(r) for r in ws.iter_rows(values_only=True)][1:]
    rows = [r for r in rows if r and r[0] is not None]
    categories = Counter(r[7] for r in rows if r[7] is not None)
    max_award = max(float(r[4]) for r in rows if isinstance(r[4], (int, float)))
    return [
        field("final_analysis_award_count", "integer", "How many awards are listed on the 'Final Analysis' sheet?", len(rows)),
        field("no_concern_count", "integer", "How many awards are labeled 'No Concern' on the 'Final Analysis' sheet?", categories.get("No Concern", 0)),
        field("contact_fast_spending_count", "integer", "How many awards are labeled 'Contact - Fast Spending' on the 'Final Analysis' sheet?", categories.get("Contact - Fast Spending", 0)),
        field("maximum_total_awarded_amount", "number", "What is the largest Total Awarded Amount on the 'Final Analysis' sheet?", max_award),
    ]


def extract_3f821c2d(task: Dict[str, Any], wb: openpyxl.Workbook) -> List[Dict[str, Any]]:
    ws = wb[wb.sheetnames[0]]
    return [
        field("target_gross_receipts_total", "integer", "What is the 2025 target Gross Receipts total?", int(ws.cell(27, 14).value)),
        field("july_2025_eom_inventory_stores", "integer", "What is the July 2025 EOM inventory level for STORES?", int(ws.cell(31, 14).value)),
        field("july_2025_eom_inventory_ecommerce", "integer", "What is the July 2025 EOM inventory level for ECOMMERCE?", int(ws.cell(32, 14).value)),
        field("planned_omni_turn_total", "number", "What is the planned OMNI total turn value?", float(ws.cell(24, 20).value)),
    ]


def extract_bf68f2ad(task: Dict[str, Any], wb: openpyxl.Workbook) -> List[Dict[str, Any]]:
    ws = wb[wb.sheetnames[0]]
    catchup = [ws.cell(2, c).value for c in range(2, 9)]
    days = ws.cell(3, 2).value
    return [
        field("week_4_catch_up_hours", "number", "What is the catch-up plan value for week 4?", float(catchup[0])),
        field("week_9_catch_up_hours", "number", "What is the catch-up plan value for week 9?", float(catchup[5])),
        field("total_catch_up_hours", "number", "What is the total catch-up plan across weeks 4 through 10?", float(sum(catchup))),
        field("days_per_week_in_plan", "integer", "How many days per week are used in the catch-up plan?", int(days)),
    ]


def extract_68d8d901(task: Dict[str, Any], wb: openpyxl.Workbook) -> List[Dict[str, Any]]:
    ws = wb["Work Schedule"]
    assignments = wb["Production Assignments"]
    target = ws.cell(1, 4).value
    shift_length = ws.cell(2, 4).value
    shifts = ws.cell(3, 4).value
    operators = assignments.cell(1, 3).value
    return [
        field("production_target_lbs", "integer", "What production target in pounds is listed on the Work Schedule?", int(str(target).split()[0].replace(",", ""))),
        field("shift_length_hours", "integer", "How many hours is each shift in the Work Schedule?", int(str(shift_length).split()[0])),
        field("shifts_per_day", "integer", "How many shifts per day are listed in the Work Schedule?", int(str(shifts).split('/')[0])),
        field("operators_per_shift", "integer", "How many employees per 12-hour shift are listed in Production Assignments?", int(str(operators).split()[0])),
    ]


def extract_5a2d70da(task: Dict[str, Any], wb: openpyxl.Workbook) -> List[Dict[str, Any]]:
    ws = wb["COVER PLATE"]
    step_count = 0
    for row in ws.iter_rows(min_row=8, values_only=True):
        if row[0] is not None:
            step_count += 1
    return [
        field("part_name", "text", "What part name is listed in the workbook?", str(ws.cell(1, 2).value).strip()),
        field("material_type", "text", "What material type is listed for the part?", str(ws.cell(2, 2).value).strip()),
        field("number_of_operations", "integer", "How many operations are listed for the part?", int(ws.cell(4, 2).value)),
        field("machining_step_count", "integer", "How many machining steps are listed in the feature table?", step_count),
    ]


def extract_81db15ff(task: Dict[str, Any], wb: openpyxl.Workbook) -> List[Dict[str, Any]]:
    ws = wb["Sheet1"]
    rows = [list(r) for r in ws.iter_rows(values_only=True)][1:]
    rows = [r for r in rows if r and r[0] not in (None, "")]
    np_yes = sum(1 for r in rows if isinstance(r[1], str) and r[1].strip().lower().startswith("yes"))
    pa_yes = sum(1 for r in rows if isinstance(r[2], str) and r[2].strip().lower().startswith("yes"))
    pa_cosign_yes = sum(1 for r in rows if isinstance(r[4], str) and r[4].strip().lower().startswith("yes"))
    return [
        field("state_count", "integer", "How many states are included in the comparison table?", len(rows)),
        field("np_independent_yes_state_count", "integer", "In how many states can NPs practice independently?", np_yes),
        field("pa_independent_yes_state_count", "integer", "In how many states can PAs practice independently?", pa_yes),
        field("pa_cosignature_required_state_count", "integer", "In how many states do PA charts require co-signature?", pa_cosign_yes),
    ]


def extract_61e7b9c6(task: Dict[str, Any], wb: openpyxl.Workbook) -> List[Dict[str, Any]]:
    from collections import Counter

    ws = wb["MHT Formulary"]
    medication_rows = 0
    group = None
    group_counts = Counter()
    max_cost = 0
    for row in ws.iter_rows(values_only=True):
        values = list(row)
        if values[0] and values[1] is None and values[2] is None:
            group = str(values[0]).strip()
            continue
        if values[1] not in (None, ""):
            medication_rows += 1
            group_counts[group] += 1
            if isinstance(values[6], (int, float)):
                max_cost = max(max_cost, float(values[6]))
    return [
        field("medication_entry_count", "integer", "How many medication entries are listed in the formulary?", medication_rows),
        field("transdermal_estradiol_entry_count", "integer", "How many entries are in the 'ESTRADIOL (TRANSDERMAL)' group?", group_counts.get("ESTRADIOL (TRANSDERMAL)", 0)),
        field("vaginal_estradiol_entry_count", "integer", "How many entries are in the 'ESTRADIOL VAGINAL' group?", group_counts.get("ESTRADIOL VAGINAL", 0)),
        field("highest_known_estimated_cost", "number", "What is the highest known estimated cost without insurance in the formulary?", max_cost),
    ]


def extract_dd724c67(task: Dict[str, Any], wb: openpyxl.Workbook) -> List[Dict[str, Any]]:
    hospitals = wb["Hospitals"]
    rehabs = wb["Rehabs, SNFs"]
    return [
        field("hospital_count", "integer", "How many hospitals are listed in the Hospitals sheet?", hospitals.max_row - 2),
        field("rehab_snf_count", "integer", "How many facilities are listed in the Rehabs, SNFs sheet?", rehabs.max_row - 2),
        field("high_acuity_follow_up_days", "integer", "How many days are recommended for high-acuity follow-up?", 7),
        field("low_acuity_follow_up_weeks", "integer", "How many weeks are recommended for low-acuity follow-up?", 6),
    ]


def extract_90edba97(task: Dict[str, Any], wb: openpyxl.Workbook) -> List[Dict[str, Any]]:
    def patient_count(sheet_name: str) -> int:
        ws = wb[sheet_name]
        header = [c for c in next(ws.iter_rows(values_only=True))]
        return sum(1 for v in header[1:] if v not in (None, ""))

    return [
        field("tracked_patient_count", "integer", "How many patients are tracked on each management sheet?", patient_count("Adequacy Tracker")),
        field("adequacy_spktv_threshold", "number", "What spKt/V threshold is listed in the Clinical Guidelines Reference?", 1.2),
        field("phosphorus_upper_threshold", "number", "What is the upper phosphorus threshold listed in the Clinical Guidelines Reference?", 5.5),
        field("uncorrected_calcium_threshold", "number", "What is the uncorrected calcium threshold listed in the Clinical Guidelines Reference?", 10.2),
    ]


def extract_f2986c1f(task: Dict[str, Any], wb: openpyxl.Workbook) -> List[Dict[str, Any]]:
    from collections import Counter

    ws = wb["Sheet1"]
    rows = [list(r) for r in ws.iter_rows(values_only=True)][1:]
    rows = [r for r in rows if r and r[0] is not None]
    types = Counter(r[6] for r in rows)
    acetaminophen_count = sum(1 for r in rows if r[4] == "Acetaminophen")
    return [
        field("medication_count", "integer", "How many medications are listed in the spreadsheet?", len(rows)),
        field("over_the_counter_count", "integer", "How many medications are labeled 'Over the counter'?", types.get("Over the counter ", 0)),
        field("legend_drug_count", "integer", "How many medications are labeled 'Legend Drug'?", types.get("Legend Drug", 0)),
        field("acetaminophen_entry_count", "integer", "How many Acetaminophen entries are listed?", acetaminophen_count),
    ]


def extract_b57efde3(task: Dict[str, Any], wb: openpyxl.Workbook) -> List[Dict[str, Any]]:
    from collections import Counter

    ws = wb["Exhibitors"]
    rows = [list(r) for r in ws.iter_rows(values_only=True)][1:]
    rows = [r for r in rows if r and r[0] is not None]
    countries = Counter(r[1] for r in rows)
    uc_count = sum(1 for r in rows if isinstance(r[3], str) and "UC" in r[3])
    rov_count = sum(1 for r in rows if isinstance(r[3], str) and "ROV" in r[3])
    return [
        field("exhibitor_count", "integer", "How many exhibitors are listed in the prospect sheet?", len(rows)),
        field("norway_company_count", "integer", "How many exhibitors are based in Norway?", countries.get("Norway", 0)),
        field("uc_company_count", "integer", "How many exhibitors list UC products?", uc_count),
        field("rov_company_count", "integer", "How many exhibitors list ROV products?", rov_count),
    ]


def extract_4520f882(task: Dict[str, Any], wb: openpyxl.Workbook) -> List[Dict[str, Any]]:
    ws = wb["Schedule"]
    schedule_rows = [list(r) for r in ws.iter_rows(min_row=4, values_only=True) if r[1] is not None and r[3] is not None]
    performance_count = sum(1 for r in schedule_rows if r[3] == "Performance")

    roster = wb["Roster"]
    musicians = [list(r) for r in roster.iter_rows(min_row=6, max_row=15, values_only=True) if r[0] is not None]
    parker = next(r for r in musicians if r[0] == "Parker")
    total_payroll = sum(float(r[28]) for r in musicians if isinstance(r[28], (int, float)))
    return [
        field("scheduled_service_count", "integer", "How many populated service rows are listed on the Schedule sheet?", len(schedule_rows)),
        field("performance_service_count", "integer", "How many Performance services are listed on the Schedule sheet?", performance_count),
        field("musician_count", "integer", "How many musicians are listed on the Roster sheet?", len(musicians)),
        field("parker_total_earnings", "number", "What is Parker's total earnings on the Roster sheet?", float(parker[28]), tolerance=0.01),
    ]


def extract_40a8c4b1(task: Dict[str, Any], wb: openpyxl.Workbook) -> List[Dict[str, Any]]:
    from collections import Counter

    ws = wb["2025 Grand Rounds Schedule"]
    rows = [r for r in ws.iter_rows(values_only=True) if r and hasattr(r[0], "year")]
    counts = Counter(r[2] for r in rows)
    return [
        field("grand_rounds_session_count", "integer", "How many total scheduled Grand Rounds sessions are listed?", len(rows)),
        field("grand_rounds_date_count", "integer", "How many unique Grand Rounds dates are scheduled?", len({r[0] for r in rows})),
        field("m_m_session_count", "integer", "How many M&M sessions are scheduled?", counts.get("M&M", 0)),
        field("all_periop_meeting_count", "integer", "How many All Periop Meeting sessions are scheduled?", counts.get("All Periop Meeting", 0)),
    ]


def extract_1752cb53(task: Dict[str, Any], wb: openpyxl.Workbook) -> List[Dict[str, Any]]:
    ws = wb["One Week Test Plan"]
    rows = [list(r) for r in ws.iter_rows(min_row=2, max_row=43, values_only=True) if r[0] is not None]
    return [
        field("planned_sku_count", "integer", "How many SKU rows are populated in the One Week Test Plan?", len(rows)),
        field("total_fg_packs_needed", "integer", "What is the total FG Packs Needed across the plan?", int(sum(r[1] for r in rows))),
        field("total_wire_yards", "integer", "What is the total yards of wire required across the plan?", int(sum(r[3] for r in rows))),
        field("total_changeover_minutes", "integer", "What is the total changeover time in minutes across the plan?", int(sum((r[5] or 0) for r in rows))),
    ]


def extract_0fad6023(task: Dict[str, Any], wb: openpyxl.Workbook) -> List[Dict[str, Any]]:
    ws = wb["FSC POG"]
    pan_slots = sum(1 for c in range(2, 200) if ws.cell(5, c).value not in (None, ""))
    return [
        field("case_size_inches", "integer", "What case size in inches is shown on the FSC POG sheet?", int(ws.cell(1, 2).value)),
        field("space_used_inches", "integer", "How many inches of case space are used on the FSC POG sheet?", int(ws.cell(2, 2).value)),
        field("space_available_inches", "integer", "How many inches of case space remain available on the FSC POG sheet?", int(ws.cell(3, 2).value)),
        field("pan_slot_count", "integer", "How many pan slots are shown in the FSC POG layout?", pan_slots),
    ]


def extract_02aa1805(task: Dict[str, Any], wb: openpyxl.Workbook) -> List[Dict[str, Any]]:
    ws = wb["Potential Wells"]
    rows = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        if r[0] is None or not isinstance(r[4], (int, float)):
            continue
        rows.append(list(r))
    max_depth = max(r[4] for r in rows)
    deepest = next(r for r in rows if r[4] == max_depth)
    return [
        field("recommended_well_count", "integer", "How many wells are listed on the Potential Wells sheet?", len(rows)),
        field("farmer_city_recommended_count", "integer", "How many recommended wells are in FARMER CITY?", sum(1 for r in rows if r[0] == "FARMER CITY")),
        field("deepest_recommended_well_id", "text", "What is the well ID of the deepest recommended well?", str(deepest[1]).strip()),
        field("zero_pumpage_recommended_count", "integer", "How many recommended wells have pumpage equal to 0?", sum(1 for r in rows if r[6] == 0)),
    ]


def extract_a99d85fc(task: Dict[str, Any], wb: openpyxl.Workbook) -> List[Dict[str, Any]]:
    ws = wb["Lease Terms"]
    return [
        field("suite_size_sf", "integer", "What suite size in square feet is used in the lease calculator?", int(ws.cell(4, 3).value)),
        field("scenario_1_gross_lease_value", "number", "What is the Gross Lease Value for Scenario 1?", float(ws.cell(15, 8).value), tolerance=0.01),
        field("scenario_2_gross_lease_value", "number", "What is the Gross Lease Value for Scenario 2?", float(ws.cell(15, 12).value), tolerance=0.01),
        field("scenario_3_gross_lease_value", "number", "What is the Gross Lease Value for Scenario 3?", float(ws.cell(15, 16).value), tolerance=0.01),
    ]


def extract_650adcb1(task: Dict[str, Any], wb: openpyxl.Workbook) -> List[Dict[str, Any]]:
    months = ["December", "January", "February", "March", "April"]
    req_total = 0
    req_counts = {"Adam Blake": 0, "Dustin Herman ": 0, "Katie Montgomery ": 0}
    work_counts = {"Adam Blake": 0, "Dustin Herman ": 0, "Katie Montgomery ": 0}
    names = {6: "Adam Blake", 7: "Dustin Herman ", 8: "Katie Montgomery "}
    for s in months:
        ws = wb[s]
        for row_idx, name in names.items():
            values = [ws.cell(row_idx, c).value for c in range(4, 35)]
            for value in values:
                if value == "x":
                    work_counts[name] += 1
                elif isinstance(value, str) and "Requested Day Off" in value:
                    req_counts[name] += 1
                    req_total += 1
    time_off = wb["Time off Requests"]
    conflict_count = sum(1 for r in range(27, 80) if time_off.cell(r, 5).value)
    return [
        field("requested_day_off_total", "integer", "How many requested day off entries are scheduled across the winter calendar?", req_total),
        field("adam_requested_day_off_count", "integer", "How many requested day off entries does Adam Blake have on the calendar?", req_counts["Adam Blake"]),
        field("katie_total_work_days", "integer", "How many working days are scheduled for Katie Montgomery across the winter calendar?", work_counts["Katie Montgomery "]),
        field("understaffed_conflict_day_count", "integer", "How many dates are listed as having fewer than two interns scheduled to work?", conflict_count),
    ]


def extract_11dcc268(task: Dict[str, Any], wb: openpyxl.Workbook) -> List[Dict[str, Any]]:
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(min_row=3, values_only=True) if r[0] is not None]
    partial_row = next(r for r in rows if r[4] == "A-300-K11")
    return [
        field("location_report_line_count", "integer", "How many populated inventory movement lines are listed in the location report?", len(rows)),
        field("total_quantity_received", "integer", "What is the total quantity received across the location report?", int(sum(r[2] or 0 for r in rows if r[2] is not None))),
        field("remaining_balance_quantity", "integer", "What is the total remaining balance quantity across the location report?", int(sum(r[6] or 0 for r in rows))),
        field("partial_receipt_moved_to_location", "text", "What location was the partial receipt moved to for item P07-P98K45-20?", str(partial_row[4]).strip()),
    ]


CUSTOM_EXTRACTORS = {
    "efca245f-c24f-4f75-a9d5-59201330ab7a": extract_efca245f,
    "83d10b06-26d1-4636-a32c-23f92c57f30b": extract_83d10b06,
    "7bbfcfe9-132d-4194-82bb-d6f29d001b01": extract_7bbfcfe9,
    "dfb4e0cd-a0b7-454e-b943-0dd586c2764c": extract_dfb4e0cd,
    "3f821c2d-ab97-46ec-a0fb-b8f73c2682bc": extract_3f821c2d,
    "bf68f2ad-eac5-490a-adec-d847eb45bd6f": extract_bf68f2ad,
    "68d8d901-dd0b-4a7e-bf9a-1074fddf1a96": extract_68d8d901,
    "5a2d70da-0a42-4a6b-a3ca-763e03f070a5": extract_5a2d70da,
    "81db15ff-ceea-4f63-a1cd-06dc88114709": extract_81db15ff,
    "61e7b9c6-0051-429f-a341-fda9b6578a84": extract_61e7b9c6,
    "dd724c67-8118-4b99-ab50-4761af705c3b": extract_dd724c67,
    "90edba97-74f0-425a-8ff6-8b93182eb7cb": extract_90edba97,
    "f2986c1f-2bbf-4b83-bc93-624a9d617f45": extract_f2986c1f,
    "b57efde3-26d6-4742-bbff-2b63c43b4baa": extract_b57efde3,
    "4520f882-715a-482d-8e87-1cb3cbdfe975": extract_4520f882,
    "40a8c4b1-b169-4f92-a38b-7f79685037ec": extract_40a8c4b1,
    "1752cb53-5983-46b6-92ee-58ac85a11283": extract_1752cb53,
    "0fad6023-767b-42c1-a1b3-027cd4f583cb": extract_0fad6023,
    "02aa1805-c658-4069-8a6a-02dec146063a": extract_02aa1805,
    "a99d85fc-eff8-48d2-a7d4-42a75d62f18d": extract_a99d85fc,
    "650adcb1-ed19-4f88-8117-77640f7b94b6": extract_650adcb1,
    "11dcc268-cb07-4d3a-a184-c6d7a19349bc": extract_11dcc268,
}


def fallback_extract(wb: openpyxl.Workbook) -> List[Dict[str, Any]]:
    """Deterministic extraction from labeled totals/averages/counts."""
    candidates: List[Tuple[str, str, float]] = []
    label_re = re.compile(r"(total|grand total|average|avg|count|net|sum|subtotal)", re.IGNORECASE)

    for sname in wb.sheetnames[:4]:
        ws = wb[sname]
        for row in ws.iter_rows(min_row=1, max_row=400, min_col=1, max_col=30, values_only=True):
            row_vals = list(row)
            text_cells = [str(c).strip() for c in row_vals if isinstance(c, str) and str(c).strip()]
            num_cells = [to_number(c) for c in row_vals]
            num_cells = [n for n in num_cells if n is not None]
            if not text_cells or not num_cells:
                continue
            label = text_cells[0]
            if not label_re.search(label):
                continue
            # Use rightmost numeric value in the row as likely final metric
            value = num_cells[-1]
            if value is None:
                continue
            candidates.append((sname, label, value))

    # Deduplicate by normalized label
    seen = set()
    fields: List[Dict[str, Any]] = []
    for sname, label, value in candidates:
        key = snake(f"{sname}_{label}")[:60]
        if key in seen:
            continue
        seen.add(key)

        v = round_value(value)
        is_int = isinstance(v, int)
        ftype = "integer" if is_int else "number"
        field = {
            "key": key,
            "type": ftype,
            "description": f"What is '{label}' in sheet '{sname}'?",
            "expected": v,
        }
        if ftype == "number":
            field["tolerance"] = tolerance_for_value(float(v))
        fields.append(field)
        if len(fields) >= 4:
            break

    # Ensure minimum 2 fields
    if len(fields) < 2:
        first = wb[wb.sheetnames[0]]
        non_empty_rows = 0
        for row in first.iter_rows(values_only=True):
            if any(c is not None and str(c).strip() != "" for c in row):
                non_empty_rows += 1

        if not any(f["key"] == "sheet_count" for f in fields):
            fields.append(
                {
                    "key": "sheet_count",
                    "type": "integer",
                    "description": "How many sheets does the workbook have?",
                    "expected": len(wb.sheetnames),
                }
            )
        if len(fields) < 2:
            fields.append(
                {
                    "key": f"data_rows_{snake(wb.sheetnames[0])}",
                    "type": "integer",
                    "description": f"How many non-empty rows are in sheet '{wb.sheetnames[0]}'?",
                    "expected": non_empty_rows,
                }
            )

    return fields[:4]


def llm_extract(
    client: Any,
    model: str,
    task: Dict[str, Any],
    wb: openpyxl.Workbook,
) -> List[Dict[str, Any]]:
    samples = {}
    for sname in wb.sheetnames[:3]:
        samples[sname] = sample_sheet(wb[sname], max_rows=10, max_cols=10)

    prompt = (
        "You are analyzing an Excel deliverable from a data analysis task.\n\n"
        f"Task description:\n{task.get('prompt', '')[:800]}\n\n"
        f"Sheets: {wb.sheetnames[:8]}\n"
        f"Samples: {json.dumps(samples, ensure_ascii=False)}\n\n"
        "Identify 2-4 key FINAL answer values to verify correctness. "
        "Prefer totals, counts, averages, or final metrics over intermediate values.\n"
        "Return JSON with key 'fields' only:\n"
        "{\"fields\": ["
        "{\"key\":\"...\",\"type\":\"number|integer|text\","
        "\"description\":\"...\",\"value\":123.4}]}"
    )

    resp = client.chat.completions.create(
        model=model,
        messages=[{"role": "user", "content": prompt}],
        temperature=0.1,
        response_format={"type": "json_object"},
    )
    content = resp.choices[0].message.content
    parsed = json.loads(content)
    raw_fields = parsed.get("fields", [])

    fields: List[Dict[str, Any]] = []
    for rf in raw_fields[:4]:
        key = snake(str(rf.get("key", "value")))[:60]
        ftype = str(rf.get("type", "number")).lower()
        if ftype not in {"number", "integer", "text"}:
            ftype = "number"
        desc = str(rf.get("description", "")) or f"What is {key}?"
        val = rf.get("value")
        if ftype in {"number", "integer"}:
            n = to_number(val)
            if n is None:
                continue
            if ftype == "integer":
                val = int(round(n))
            else:
                val = round_value(float(n))

        field = {
            "key": key,
            "type": ftype,
            "description": desc,
            "expected": val,
        }
        if ftype == "number":
            field["tolerance"] = tolerance_for_value(float(val))
        fields.append(field)

    return fields[:4]


def select_tasks(tasks: List[Dict[str, Any]], mode: str) -> List[Dict[str, Any]]:
    if mode == "missing":
        return [t for t in tasks if not t.get("submission_fields")]
    if mode == "generic":
        return [t for t in tasks if is_generic_fields(t.get("submission_fields", []))]
    return tasks


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate submission_fields with optional LLM assistance")
    parser.add_argument("--manifest", default=str(DEFAULT_MANIFEST))
    parser.add_argument("--model", default="Qwen/Qwen3-235B-A22B-Instruct-2507")
    parser.add_argument("--base-url", default="https://router.huggingface.co/v1")
    parser.add_argument("--api-key-env", default="HF_TOKEN")
    parser.add_argument("--mode", choices=["all", "missing", "generic"], default="generic")
    parser.add_argument("--max-tasks", type=int, default=0)
    parser.add_argument("--use-llm", action="store_true", help="Enable optional LLM extraction (disabled by default)")
    args = parser.parse_args()

    manifest_path = Path(args.manifest)
    tasks = json.loads(manifest_path.read_text(encoding="utf-8"))
    to_process = select_tasks(tasks, args.mode)
    if args.max_tasks > 0:
        to_process = to_process[: args.max_tasks]

    client = None
    if args.use_llm:
        api_key = os.environ.get(args.api_key_env)
        if api_key:
            try:
                from openai import OpenAI

                client = OpenAI(api_key=api_key, base_url=args.base_url)
            except Exception:
                client = None

    print(f"Loaded tasks: {len(tasks)}")
    print(f"Processing ({args.mode}): {len(to_process)}")
    print(f"LLM enabled: {'yes' if client else 'no'}")

    processed = 0
    for task in to_process:
        task_id = task["task_id"]
        deliverables = task.get("expected_deliverables", [])
        if not deliverables:
            continue

        gold_path = WORKSPACES_DIR / task_id / "gold" / deliverables[0]
        if not gold_path.exists() or gold_path.suffix.lower() != ".xlsx":
            continue

        print(f"\n[{task_id[:8]}] {gold_path.name}")
        try:
            wb = openpyxl.load_workbook(str(gold_path), data_only=True, read_only=True)
            fields = []
            if task_id in CUSTOM_EXTRACTORS:
                fields = CUSTOM_EXTRACTORS[task_id](task, wb)
                print(f"  Custom fields: {len(fields)}")
            elif client:
                try:
                    fields = llm_extract(client, args.model, task, wb)
                    print(f"  LLM fields: {len(fields)}")
                except Exception as e:
                    print(f"  LLM failed: {e}")
                    fields = []

            if len(fields) < 2:
                fields = fallback_extract(wb)
                print(f"  Fallback fields: {len(fields)}")

            wb.close()

            # Final normalization
            norm_fields = []
            seen_keys = set()
            for f in fields[:4]:
                key = snake(str(f.get("key", "value")))[:60]
                if key in seen_keys:
                    continue
                seen_keys.add(key)
                ftype = str(f.get("type", "number")).lower()
                if ftype not in {"number", "integer", "text"}:
                    ftype = "number"
                expected = f.get("expected")
                if ftype in {"number", "integer"}:
                    n = to_number(expected)
                    if n is None:
                        continue
                    expected = int(round(n)) if ftype == "integer" else round_value(float(n))
                out = {
                    "key": key,
                    "type": ftype,
                    "description": str(f.get("description", f"What is {key}?")),
                    "expected": expected,
                }
                if ftype == "number":
                    out["tolerance"] = float(f.get("tolerance", tolerance_for_value(float(expected))))
                norm_fields.append(out)

            if len(norm_fields) >= 2:
                task["submission_fields"] = norm_fields[:4]
                processed += 1
                for fld in norm_fields[:4]:
                    print(f"    - {fld['key']} = {fld['expected']}")
            else:
                print("  skipped: could not produce at least 2 valid fields")

        except Exception as e:
            print(f"  error: {e}")
            continue

    manifest_path.write_text(json.dumps(tasks, indent=2, ensure_ascii=False), encoding="utf-8")
    print("\n" + "=" * 70)
    print(f"Updated manifest: {manifest_path}")
    print(f"Tasks updated this run: {processed}")


if __name__ == "__main__":
    main()

