"""
Reward computation for FinBench.

Three-signal rubric architecture:
  Signal 1: Structural checks  (0.25 of rubric_score)
  Signal 2: Submission fields   (0.60 of rubric_score)
  Signal 3: File-vs-summary     (0.15 of rubric_score)

Plus trace-based scores:
  Execution quality  — 0.25 weight
  Memory & process   — 0.25 weight
  Uncapped depth bonus — 0.10 weight

Overall:
  total = 0.50 * rubric + 0.25 * exec + 0.25 * memory + 0.10 * depth
"""

import os
import re
from typing import Any, Dict, List, Optional, Set, Tuple

from .trace import TraceLogger
from .workspace import Workspace


# ═══════════════════════════════════════════════════
# Top-level entry point
# ═══════════════════════════════════════════════════

def compute_total_reward(
    task: Dict,
    trace: TraceLogger,
    workspace: Workspace,
    executor: Any,
    workspace_root: str,
    submission_values: Dict[str, Any] = None,
) -> Dict[str, Any]:
    """Compute the full reward breakdown using the 3-signal rubric."""
    deliverable_path = _resolve_deliverable(task, workspace_root)

    structural, structural_details = compute_structural_score(
        task, workspace_root, deliverable_path,
    )
    submission, submission_details = compute_submission_score(
        task, submission_values,
    )
    consistency, consistency_details = compute_consistency_score(
        task, submission_values, deliverable_path,
    )

    rubric_score = 0.25 * structural + 0.60 * submission + 0.15 * consistency

    exec_quality, exec_checks = compute_execution_quality(trace, task, workspace_root)
    memory_process, memory_checks = compute_memory_process_score(
        trace, workspace, workspace_root, executor,
    )
    depth, depth_details = compute_depth_bonus(trace, workspace, workspace_root, executor)

    total = (
        0.50 * rubric_score
        + 0.25 * exec_quality
        + 0.25 * memory_process
        + 0.10 * depth
    )

    return {
        "total_reward": round(total, 4),
        "rubric_score": round(rubric_score, 4),
        "structural_score": round(structural, 4),
        "submission_score": round(submission, 4),
        "consistency_score": round(consistency, 4),
        "execution_quality": round(exec_quality, 4),
        "memory_process": round(memory_process, 4),
        "depth_bonus": round(depth, 4),
        "structural_details": structural_details,
        "submission_details": submission_details,
        "consistency_details": consistency_details,
        "execution_checks": exec_checks,
        "memory_checks": memory_checks,
        "depth_details": depth_details,
    }


def _resolve_deliverable(task: Dict, workspace_root: str) -> Optional[str]:
    deliverables = task.get("expected_deliverables", [])
    if deliverables:
        return os.path.join(workspace_root, "output", deliverables[0])
    return None


# ═══════════════════════════════════════════════════
# Signal 1: Structural Checks (0.25 weight)
# ═══════════════════════════════════════════════════

def compute_structural_score(
    task: Dict,
    workspace_root: str,
    deliverable_path: Optional[str],
) -> Tuple[float, List[Dict]]:
    """Auto-parse rubric NL text into deterministic structural checks."""
    raw_rubric = task.get("rubric", [])
    if not raw_rubric:
        return 0.0, []

    if isinstance(raw_rubric, dict):
        structural_items = raw_rubric.get("structural") or []
    else:
        structural_items = raw_rubric

    if not structural_items:
        return 0.0, []

    checks: List[Dict] = []
    for item in structural_items:
        if isinstance(item, str):
            criterion = item
            score = 1
        else:
            criterion = item.get("criterion", "")
            score = item.get("score", 1)
        checks.extend(_parse_structural_checks(criterion, score, deliverable_path))

    if not checks:
        return 0.0, []

    total_weight = sum(c["weight"] for c in checks)
    earned = 0.0
    details = []

    for check in checks:
        passed, reason = _run_structural_check(check, workspace_root)
        if passed:
            earned += check["weight"]
        details.append({
            "type": check["type"],
            "weight": check["weight"],
            "passed": passed,
            "reason": reason,
            "spec": {k: v for k, v in check.items() if k not in ("type", "weight")},
        })

    return (earned / total_weight if total_weight > 0 else 0.0), details


def _parse_structural_checks(
    criterion: str, score: int, deliverable_path: Optional[str],
) -> List[Dict]:
    """Regex-match easy patterns from a single rubric criterion string."""
    c = criterion.lower()
    found: List[Dict] = []
    file_path = deliverable_path or ""

    if re.search(r"\b(xlsx|excel\s+workbook|single\s+file)\b", c):
        found.append({
            "type": "file_exists_xlsx",
            "file": file_path,
            "weight": score,
        })

    header_match = re.search(
        r"first\s+row\s+(?:includes?|contains?|has)\s+(.+)",
        c,
    )
    if header_match:
        raw = header_match.group(1)
        headers = [h.strip().strip("'\"") for h in re.split(r",\s*(?:and\s+)?|;\s*", raw) if h.strip()]
        if headers:
            found.append({
                "type": "sheet_has_headers",
                "file": file_path,
                "headers": headers,
                "weight": score,
            })

    range_match = re.search(r"between\s+(\d+)\s+and\s+(\d+)", c)
    if range_match and re.search(r"\b(rows?|count|test\s+cases?)\b", c):
        lo, hi = int(range_match.group(1)), int(range_match.group(2))
        found.append({
            "type": "row_count_range",
            "file": file_path,
            "min_rows": lo,
            "max_rows": hi,
            "weight": score,
        })

    blank_match = re.search(r"\b(blank|empty)\b.*?['\"]([^'\"]+)['\"]", c)
    if not blank_match:
        blank_match = re.search(r"['\"]([^'\"]+)['\"]\s+.*?\b(blank|empty)\b", c)
        if blank_match:
            col_name = blank_match.group(1)
            found.append({
                "type": "column_all_blank",
                "file": file_path,
                "column": col_name,
                "weight": score,
            })
    else:
        col_name = blank_match.group(2)
        found.append({
            "type": "column_all_blank",
            "file": file_path,
            "column": col_name,
            "weight": score,
        })

    filled_match = re.search(r"\b(non[- ]?empty|non[- ]?blank|filled)\b.*?['\"]([^'\"]+)['\"]", c)
    if not filled_match:
        filled_match = re.search(r"['\"]([^'\"]+)['\"]\s+.*?\b(non[- ]?empty|non[- ]?blank|filled)\b", c)
        if filled_match:
            col_name = filled_match.group(1)
            found.append({
                "type": "column_all_filled",
                "file": file_path,
                "column": col_name,
                "weight": score,
            })
    else:
        col_name = filled_match.group(2)
        found.append({
            "type": "column_all_filled",
            "file": file_path,
            "column": col_name,
            "weight": score,
        })

    return found


def _run_structural_check(check: Dict, workspace_root: str) -> Tuple[bool, str]:
    """Execute a single structural check."""
    ctype = check["type"]
    file_path = check.get("file", "")

    if ctype == "file_exists_xlsx":
        exists = os.path.isfile(file_path) and file_path.lower().endswith(".xlsx")
        return exists, f"{'found' if exists else 'missing'}: {os.path.basename(file_path)}"

    if ctype == "sheet_has_headers":
        return _check_sheet_has_headers(file_path, check.get("headers", []))

    if ctype == "row_count_range":
        return _check_row_count_range(
            file_path, check.get("min_rows", 0), check.get("max_rows", 999999),
        )

    if ctype == "column_all_blank":
        return _check_column_blank(file_path, check.get("column", ""), expect_blank=True)

    if ctype == "column_all_filled":
        return _check_column_blank(file_path, check.get("column", ""), expect_blank=False)

    return False, f"unknown structural check: {ctype}"


def _check_sheet_has_headers(
    xlsx_path: str, expected_headers: List[str],
) -> Tuple[bool, str]:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        ws = wb.worksheets[0]
        first_row = [
            str(c).strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            if c is not None
        ]
        wb.close()
    except Exception as e:
        return False, f"cannot read headers: {e}"

    missing = [h for h in expected_headers if h.strip().lower() not in first_row]
    if missing:
        return False, f"missing headers: {missing}; found: {first_row}"
    return True, f"all {len(expected_headers)} headers present"


def _check_row_count_range(
    xlsx_path: str, min_rows: int, max_rows: int,
) -> Tuple[bool, str]:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        ws = wb.worksheets[0]
        count = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True))
        wb.close()
    except Exception as e:
        return False, f"cannot count rows: {e}"

    passed = min_rows <= count <= max_rows
    return passed, f"row count {count} (expected {min_rows}–{max_rows})"


def _check_column_blank(
    xlsx_path: str, column_name: str, expect_blank: bool,
) -> Tuple[bool, str]:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        ws = wb.worksheets[0]

        headers = [
            str(c).strip().lower() if c else ""
            for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        ]
        col_lower = column_name.strip().lower()
        if col_lower not in headers:
            wb.close()
            return False, f"column '{column_name}' not found; have {headers}"
        col_idx = headers.index(col_lower)

        values = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if col_idx < len(row):
                values.append(row[col_idx])

        wb.close()
    except Exception as e:
        return False, f"cannot read column: {e}"

    if expect_blank:
        all_blank = all(v is None or str(v).strip() == "" for v in values)
        return all_blank, f"column '{column_name}': {'all blank' if all_blank else 'has non-blank values'}"
    else:
        all_filled = all(v is not None and str(v).strip() != "" for v in values)
        return all_filled, f"column '{column_name}': {'all filled' if all_filled else 'has blank values'}"


# ═══════════════════════════════════════════════════
# Signal 2: Submission Field Score (0.60 weight)
# ═══════════════════════════════════════════════════

def compute_submission_score(
    task: Dict,
    submission_values: Optional[Dict[str, Any]],
) -> Tuple[float, List[Dict]]:
    """Compare submitted values against expected submission_fields."""
    fields = task.get("submission_fields", [])
    if not fields or not submission_values:
        return 0.0, []

    matched = 0
    details = []

    for field in fields:
        key = field.get("key", field.get("name", ""))
        expected = field.get("expected")
        ftype = field.get("type", "text")
        tolerance = field.get("tolerance", 1.0)

        submitted = submission_values.get(key)
        if submitted is None:
            details.append({
                "field": key,
                "passed": False,
                "reason": "not submitted",
                "expected": expected,
                "submitted": None,
            })
            continue

        passed, reason = _compare_field(submitted, expected, ftype, tolerance)
        if passed:
            matched += 1
        details.append({
            "field": key,
            "passed": passed,
            "reason": reason,
            "expected": expected,
            "submitted": submitted,
        })

    score = matched / len(fields) if fields else 0.0
    return score, details


def _compare_field(
    submitted: Any, expected: Any, ftype: str, tolerance: float,
) -> Tuple[bool, str]:
    if ftype == "number":
        try:
            s, e = float(submitted), float(expected)
            passed = abs(s - e) <= tolerance
            return passed, f"{s} vs {e} (tol={tolerance})"
        except (ValueError, TypeError):
            return False, f"cannot parse as number: submitted={submitted!r}"

    if ftype == "integer":
        try:
            s, e = int(submitted), int(expected)
            passed = s == e
            return passed, f"{s} vs {e}"
        except (ValueError, TypeError):
            return False, f"cannot parse as integer: submitted={submitted!r}"

    s_str = str(submitted).strip().lower()
    e_str = str(expected).strip().lower()
    passed = s_str == e_str
    return passed, f"'{s_str}' vs '{e_str}'"


# ═══════════════════════════════════════════════════
# Signal 3: File-vs-Summary Consistency (0.15 weight)
# ═══════════════════════════════════════════════════

def compute_consistency_score(
    task: Dict,
    submission_values: Optional[Dict[str, Any]],
    deliverable_path: Optional[str],
) -> Tuple[float, List[Dict]]:
    """Check that submitted numeric values actually appear in the output file."""
    fields = task.get("submission_fields", [])
    if not fields or not submission_values or not deliverable_path:
        return 0.0, []

    numeric_fields = [
        f for f in fields if f.get("type") in ("number", "integer")
    ]
    if not numeric_fields:
        return 1.0, [{"note": "no numeric fields to verify"}]

    file_nums = _extract_all_numerics(deliverable_path)
    if not file_nums:
        return 0.0, [{"note": f"could not extract numerics from {os.path.basename(deliverable_path or '')}"}]

    matched = 0
    details = []
    tolerance = 1.0

    for field in numeric_fields:
        key = field.get("key", field.get("name", ""))
        submitted = submission_values.get(key)
        if submitted is None:
            details.append({"field": key, "passed": False, "reason": "not submitted"})
            continue

        try:
            val = float(submitted)
        except (ValueError, TypeError):
            details.append({"field": key, "passed": False, "reason": f"non-numeric: {submitted!r}"})
            continue

        found = any(abs(val - fv) <= tolerance for fv in file_nums)
        if found:
            matched += 1
        details.append({
            "field": key,
            "passed": found,
            "reason": f"{'found' if found else 'not found'} in file (val={val})",
        })

    score = matched / len(numeric_fields) if numeric_fields else 0.0
    return score, details


def _extract_all_numerics(xlsx_path: str) -> Set[float]:
    """Read all numeric cell values from an Excel file."""
    nums: Set[float] = set()
    if not xlsx_path or not os.path.isfile(xlsx_path):
        return nums
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is None:
                        continue
                    if isinstance(cell, (int, float)):
                        nums.add(float(cell))
                    else:
                        try:
                            nums.add(float(str(cell).replace(",", "")))
                        except (ValueError, TypeError):
                            pass
        wb.close()
    except Exception:
        pass
    return nums


# ═══════════════════════════════════════════════════
# Tier 2: Execution Quality (unchanged)
# ═══════════════════════════════════════════════════

def compute_execution_quality(trace: TraceLogger, task: Dict, workspace_root: str) -> Tuple[float, Dict]:
    """8 binary checks from the trace."""
    checks = {
        "explored_workspace": any(
            s["tool"] == "list_files" for s in trace.steps[:3]
        ),
        "read_reference_first": (
            trace.first_index(tool="read_file", path_contains="reference")
            < trace.first_index(tool="run_cell")
            and trace.first_index(tool="read_file", path_contains="reference") < 999
        ),
        "created_notebook": trace.count_tool("create_notebook") > 0,
        "used_search": trace.count_tool("search_files") > 0,
        "substantial_analysis": trace.count_successful_cells() >= 3,
        "error_recovery": trace.has_error_then_fix(),
        "verified_output": (
            trace.last_index(tool_in=["run_cell", "write_and_run"])
            > trace.last_index(tool="write_file", path_contains="output")
            and trace.last_index(tool="write_file", path_contains="output") >= 0
        ),
        "produced_deliverable": any(
            os.path.exists(os.path.join(workspace_root, "output", d))
            for d in task.get("expected_deliverables", ["any.xlsx"])
        ),
    }
    score = sum(1 for v in checks.values() if v) / len(checks)
    return score, checks


# ═══════════════════════════════════════════════════
# Tier 3: Memory & Process (unchanged)
# ═══════════════════════════════════════════════════

def compute_memory_process_score(
    trace: TraceLogger, workspace: Workspace, workspace_root: str, executor: Any,
) -> Tuple[float, Dict]:
    """8 binary checks for memory/process quality."""
    notebooks = _get_all_notebooks(workspace_root)

    checks = {
        "read_templates": any(
            s["tool"] == "load_from_memory"
            or (s["tool"] == "read_notebook" and "memory/" in str(s.get("args", {}).get("path", "")))
            for s in trace.steps
        ),
        "self_referenced": sum(
            1 for s in trace.steps
            if s["tool"] == "read_notebook" and "memory/" not in str(s.get("args", {}).get("path", ""))
        ) >= 1,
        "organized_workspace": trace.count_tool("create_folder") > 0,
        "documented_work": any(
            _has_markdown_cells(nb_path) for nb_path in notebooks
        ),
        "structured_notebooks": any(
            _count_cells(nb_path) >= 3 for nb_path in notebooks
        ),
        "created_intermediates": any(
            not f.startswith("reference/") and not f.startswith("memory/")
            and not f.startswith("output/") and not f.endswith(".ipynb")
            for f in workspace.get_all_files()
        ),
        "saved_to_memory": trace.count_tool("save_to_memory") > 0,
        "checked_state": trace.count_tool("get_kernel_state") > 0,
    }
    score = sum(1 for v in checks.values() if v) / len(checks)
    return score, checks


# ═══════════════════════════════════════════════════
# Uncapped Depth Bonus (unchanged)
# ═══════════════════════════════════════════════════

def compute_depth_bonus(
    trace: TraceLogger, workspace: Workspace, workspace_root: str, executor: Any,
) -> Tuple[float, Dict]:
    bonus = 0.0
    details: Dict[str, Any] = {}

    files_read = set(
        s["args"].get("path", "")
        for s in trace.steps
        if s["tool"] == "read_file" and "reference/" in str(s["args"].get("path", ""))
    )
    details["reference_files_read"] = len(files_read)
    bonus += len(files_read) * 0.1

    notebooks = _get_all_notebooks(workspace_root)
    deep_notebooks = sum(1 for nb in notebooks if _count_executed_cells(nb) >= 5)
    details["deep_notebooks"] = deep_notebooks
    bonus += deep_notebooks * 0.1

    edits = trace.count_tool("edit_cell")
    details["cell_edits"] = edits
    bonus += min(edits, 5) * 0.1

    intermediates = sum(
        1 for f in workspace.get_all_files()
        if not f.startswith("reference/") and not f.startswith("memory/")
        and not f.startswith("output/") and not f.endswith(".ipynb")
    )
    details["intermediate_files"] = intermediates
    bonus += min(intermediates, 3) * 0.1

    saved = trace.count_tool("save_to_memory") > 0
    details["saved_to_memory"] = saved
    if saved:
        bonus += 0.2

    return bonus, details


# ═══════════════════════════════════════════════════
# Utility functions (unchanged)
# ═══════════════════════════════════════════════════

def _get_all_notebooks(workspace_root: str) -> List[str]:
    notebooks = []
    for dirpath, _, filenames in os.walk(workspace_root):
        for f in filenames:
            if f.endswith(".ipynb") and "memory/" not in dirpath:
                notebooks.append(os.path.join(dirpath, f))
    return notebooks


def _has_markdown_cells(nb_path: str) -> bool:
    try:
        import nbformat
        with open(nb_path) as f:
            nb = nbformat.read(f, as_version=4)
        return any(c.cell_type == "markdown" for c in nb.cells)
    except Exception:
        return False


def _count_cells(nb_path: str) -> int:
    try:
        import nbformat
        with open(nb_path) as f:
            nb = nbformat.read(f, as_version=4)
        return len(nb.cells)
    except Exception:
        return 0


def _count_executed_cells(nb_path: str) -> int:
    try:
        import nbformat
        with open(nb_path) as f:
            nb = nbformat.read(f, as_version=4)
        return sum(1 for c in nb.cells if c.cell_type == "code" and c.get("execution_count"))
    except Exception:
        return 0
